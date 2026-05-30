"""Ledger ops + role-selector + Excel dashboard (R-004)."""
from pathlib import Path

import pytest

from orchestrator import selector
from orchestrator.dashboard import build_dashboard
from orchestrator.state import StateStore


@pytest.fixture
def store(tmp_path):
    return StateStore(state_file=tmp_path / "s.json", events_file=tmp_path / "e.jsonl",
                      active_competition="nemotron")


def _patch(*ops, tick="T1"):
    return {"tick_id": tick, "operations": list(ops)}


def op(opname, idem, **data):
    return {"op": opname, "idempotency_key": idem, "data": data}


# ---- ledger ops ----
def test_create_task_and_status_lifecycle(store):
    store.apply_patch(_patch(
        op("create_task", "k1", id="TASK-R1", title="Taxonomy", status="BACKLOG",
           allowed_area="data/taxonomy", hypothesis="H-001"),
    ))
    assert store.state["tasks"]["TASK-R1"]["title"] == "Taxonomy"
    assert store.state["tasks"]["TASK-R1"]["status"] == "BACKLOG"
    store.apply_patch(_patch(op("set_status", "k2", collection="tasks", id="TASK-R1", status="READY"), tick="T2"))
    assert store.state["tasks"]["TASK-R1"]["status"] == "READY"


def test_hypothesis_experiment_decision_metric(store):
    store.apply_patch(_patch(
        op("create_hypothesis", "h1", id="H-001", statement="data quality is the lever", status="proposed"),
        op("create_experiment", "e1", id="E-001", hypothesis="H-001", backend="kaggle_gpu", status="QUEUED"),
        op("create_decision", "d1", id="D-001", decision="adopt winner recipe", rationale="LB 0.85"),
        op("add_metric", "m1", name="cv", value=0.71, unit="acc"),
        op("update_entity", "u1", collection="experiments", id="E-001", cv_score=0.71, status="COMPLETED"),
    ))
    assert store.state["hypotheses"]["H-001"]["status"] == "proposed"
    assert store.state["experiments"]["E-001"]["cv_score"] == 0.71
    assert store.state["decisions"]["D-001"]["decision"].startswith("adopt")
    m = store.state["metrics"]
    if isinstance(m, list):
        assert m[0]["value"] == 0.71
    else:
        assert m["cv"]["value"] == 0.71


def test_ledger_replay(store):
    store.apply_patch(_patch(
        op("create_task", "k1", id="T1", title="x", status="BACKLOG"),
        op("create_hypothesis", "h1", id="H-1", statement="s", status="proposed"),
    ))
    rebuilt = store.rebuild_state_from_events()
    assert rebuilt["tasks"]["T1"]["title"] == "x"
    assert rebuilt["hypotheses"]["H-1"]["statement"] == "s"


# ---- selector ----
def _state_with(tasks=None, suggestions=None, hypotheses=None):
    return {"tasks": tasks or {}, "suggestions": suggestions or {}, "hypotheses": hypotheses or {},
            "sessions": {}, "cursors": {}}


def test_selector_picks_owner_when_ready_tasks_and_slots():
    st = _state_with(tasks={"T1": {"status": "READY"}, "T2": {"status": "READY"}})
    out = selector.select(st, free_slots=5, concurrency_target=5)
    assert out["selected"] == "owner"


def test_selector_picks_reviewer_when_prs_open():
    st = _state_with(tasks={"T1": {"status": "READY_FOR_REVIEW"}})
    out = selector.select(st, free_slots=0, open_prs=2)
    assert out["selected"] == "reviewer"


def test_selector_validator_when_validation_required():
    st = _state_with(suggestions={"S1": {"status": "VALIDATION_REQUIRED"}})
    out = selector.select(st, free_slots=0)
    assert "validator" in [c["role"] for c in out["ranked"]]


def test_selector_falls_back_to_reporter_or_planner_when_idle():
    st = _state_with()  # nothing
    out = selector.select(st, free_slots=0, concurrency_target=5)
    # planner/innovator/reporter always available; never crashes
    assert out["selected"] in ("planner", "innovator", "reporter")


def test_selector_no_adjacent_repeat_for_planner():
    st = _state_with(suggestions={"S1": {"status": "PROPOSED"}})
    st["cursors"]["last_role"] = "planner"
    out = selector.select(st, free_slots=0, previous_role="planner")
    assert out["selected"] != "planner"


# ---- dashboard ----
def test_dashboard_builds_xlsx(tmp_path, store):
    store.apply_patch(_patch(
        op("create_task", "k1", id="TASK-R1", title="Taxonomy", status="READY", hypothesis="H-001"),
        op("create_hypothesis", "h1", id="H-001", statement="data quality", status="proposed"),
    ))
    out = build_dashboard(store.state, tmp_path / "dash.xlsx")
    assert out.exists()
    from openpyxl import load_workbook
    wb = load_workbook(out)
    assert "Tasks" in wb.sheetnames and "Hypotheses" in wb.sheetnames and "Status" in wb.sheetnames
    rows = list(wb["Tasks"].iter_rows(values_only=True))
    assert any("TASK-R1" in (str(c) for c in r) for r in rows)
