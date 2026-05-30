import os
import openpyxl

def test_xlsx_exists():
    path = "competitions/nvidia-nemotron-model-reasoning-challenge/analysis/winner_code_analysis.xlsx"
    assert os.path.exists(path), "The Excel file must exist."

def test_sheets_exist():
    path = "competitions/nvidia-nemotron-model-reasoning-challenge/analysis/winner_code_analysis.xlsx"
    wb = openpyxl.load_workbook(path)
    expected_sheets = ["Overview", "TechnicalDetails", "KeyImprovements", "CoreApproach", "OpenQuestions"]
    for sheet in expected_sheets:
        assert sheet in wb.sheetnames, f"Sheet '{sheet}' is missing."

def test_overview_rows():
    path = "competitions/nvidia-nemotron-model-reasoning-challenge/analysis/winner_code_analysis.xlsx"
    wb = openpyxl.load_workbook(path)
    sheet = wb["Overview"]
    # Header + >= 17 rows
    assert sheet.max_row >= 18, "Overview sheet must have at least 17 rows (excluding header)."

def test_technical_details_rows():
    path = "competitions/nvidia-nemotron-model-reasoning-challenge/analysis/winner_code_analysis.xlsx"
    wb = openpyxl.load_workbook(path)
    sheet = wb["TechnicalDetails"]
    # Header + >= 30 rows
    assert sheet.max_row >= 31, "TechnicalDetails sheet must have at least 30 rows (excluding header)."

def test_key_improvements_rows():
    path = "competitions/nvidia-nemotron-model-reasoning-challenge/analysis/winner_code_analysis.xlsx"
    wb = openpyxl.load_workbook(path)
    sheet = wb["KeyImprovements"]
    # Header + >= 10 rows
    assert sheet.max_row >= 11, "KeyImprovements sheet must have at least 10 rows (excluding header)."
