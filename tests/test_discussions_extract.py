import os
import openpyxl

def test_discussions_extract_excel():
    excel_path = 'competitions/nvidia-nemotron-model-reasoning-challenge/analysis/discussions_extract.xlsx'

    # Check file exists
    assert os.path.exists(excel_path), f"File not found: {excel_path}"

    wb = openpyxl.load_workbook(excel_path)

    # Check 4 sheets exist
    expected_sheets = ['Threads', 'Insights', 'Top20Actionable', 'KnownGotchas']
    for sheet in expected_sheets:
        assert sheet in wb.sheetnames, f"Sheet '{sheet}' is missing."

    ws_threads = wb['Threads']
    # +1 for header
    assert ws_threads.max_row >= 218, f"Expected >= 217 threads (+1 header), got {ws_threads.max_row - 1}"

    ws_insights = wb['Insights']
    assert ws_insights.max_row >= 101, f"Expected >= 100 insights (+1 header), got {ws_insights.max_row - 1}"

    ws_top20 = wb['Top20Actionable']
    assert ws_top20.max_row == 21, f"Expected exactly 20 actionable items (+1 header), got {ws_top20.max_row - 1}"

    ws_gotchas = wb['KnownGotchas']
    assert ws_gotchas.max_row > 1, "Expected KnownGotchas to have at least one entry (+1 header)"

    # Check that it's not just the hardcoded stuff
    gotchas_vals = [ws_gotchas.cell(row=i, column=1).value for i in range(2, ws_gotchas.max_row + 1)]
    assert any("discussion-" in str(val) for val in gotchas_vals), "Gotchas should be dynamically extracted from discussion threads"

import os
import glob
import re
import openpyxl

def clean_text(text):
    return re.sub(r'\s+', ' ', text).strip()

def extract_code_blocks(text):
    code_blocks = re.findall(r'```(.*?)```', text, re.DOTALL)
    return [clean_text(cb) for cb in code_blocks]

def main():
    wb = openpyxl.Workbook()

    ws_threads = wb.active
    ws_threads.title = 'Threads'
    ws_threads.append([
        'file', 'topic_id', 'title_inferred', 'category',
        'word_count', 'author_count_inferred', 'confidence_rating_1to5'
    ])

    ws_insights = wb.create_sheet(title='Insights')
    ws_insights.append([
        'thread_file', 'insight_text', 'category',
        'applicability_to_our_notebook', 'expected_cv_delta_qualitative',
        'concrete_code_change_if_applicable'
    ])

    ws_top20 = wb.create_sheet(title='Top20Actionable')
    ws_top20.append([
        'thread_file', 'insight_text', 'category',
        'applicability_to_our_notebook', 'expected_cv_delta_qualitative',
        'concrete_code_change_if_applicable'
    ])

    ws_gotchas = wb.create_sheet(title='KnownGotchas')
    ws_gotchas.append(['failure_mode', 'fix'])

    files = glob.glob('competitions/nvidia-nemotron-model-reasoning-challenge/references/discussion-*.md')

    all_insights = []
    all_gotchas = []

    for fpath in files:
        fname = os.path.basename(fpath)
        m = re.search(r'discussion-(\d+)\.md', fname)
        topic_id = m.group(1) if m else "unknown"

        with open(fpath, 'r', encoding='utf-8') as f:
            text = f.read()

        lines = text.split('\n')
        title_inferred = "Unknown"
        for line in lines:
            if line.startswith('# '):
                title_inferred = line[2:].strip()
                break

        word_count = len(text.split())
        author_count = text.count('Posted ') + 1

        text_lower = text.lower()

        category = 'META'
        if 'rule' in text_lower or 'deadline' in text_lower:
            category = 'RULES'
        elif 'lora' in text_lower or 'train' in text_lower or 'sft' in text_lower or 'qlora' in text_lower:
            category = 'TRAIN_TRICK'
        elif 'gpu' in text_lower or 'oom' in text_lower or 'memory' in text_lower or 'blackwell' in text_lower or 'p100' in text_lower:
            category = 'GPU_TIP'
        elif 'eval' in text_lower or 'score' in text_lower or 'lb' in text_lower or 'cv' in text_lower or 'leaderboard' in text_lower:
            category = 'EVAL_TRICK'
        elif 'data' in text_lower or 'dataset' in text_lower or 'leak' in text_lower or 'augment' in text_lower:
            category = 'DATA_TRICK'
        elif 'submit' in text_lower or 'format' in text_lower or 'submission' in text_lower or 'zip' in text_lower or 'error' in text_lower:
            category = 'SUBMIT_TIP'
        elif 'mamba' in text_lower or 'cutlass' in text_lower or 'cuda' in text_lower:
            category = 'GOTCHA'

        confidence = min(5, max(1, word_count // 100))

        ws_threads.append([
            fname, topic_id, title_inferred, category,
            word_count, author_count, confidence
        ])

        # More robust insight extraction using context windows around critical phrases
        phrases = [
            ("lora", "TRAIN_TRICK", "MED"), ("qlora", "TRAIN_TRICK", "MED"),
            ("flashattention", "GPU_TIP", "HIGH"), ("gradient accumulation", "GPU_TIP", "HIGH"),
            ("batch size", "GPU_TIP", "MED"), ("chat template", "TRAIN_TRICK", "HIGH"),
            ("boxed", "SUBMIT_TIP", "HIGH"), ("mamba", "GOTCHA", "HIGH"),
            ("cutlass", "GOTCHA", "HIGH"), ("max_length", "GPU_TIP", "MED"),
            ("learning rate", "TRAIN_TRICK", "MED"), ("weight decay", "TRAIN_TRICK", "LOW")
        ]

        for p, c, cv_delta in phrases:
            if p in text_lower:
                # Find paragraphs containing the phrase
                paragraphs = re.split(r'\n\s*\n', text)
                for para in paragraphs:
                    if p in para.lower():
                        cleaned_para = clean_text(para)
                        if len(cleaned_para) > 50:
                            # Try to find a code block near here to serve as concrete code change
                            code_change = "None"
                            cb = extract_code_blocks(para)
                            if cb:
                                code_change = cb[0][:500]
                            else:
                                code_change = f"Look for code related to {p} in our notebook and verify settings."

                            all_insights.append({
                                'thread_file': fname,
                                'insight_text': cleaned_para[:500],
                                'category': c,
                                'applicability_to_our_notebook': 'MAYBE' if '?' in para else 'YES',
                                'expected_cv_delta_qualitative': cv_delta,
                                'concrete_code_change_if_applicable': code_change
                            })

                            # Also collect gotchas
                            if c == "GOTCHA" or "error" in para.lower() or "fail" in para.lower() or "oom" in para.lower() or "cuda" in para.lower():
                                all_gotchas.append({
                                    'failure_mode': f"Issue related to {p} in {fname}",
                                    'fix': cleaned_para[:300]
                                })

    # Remove duplicates
    unique_insights = {ins['insight_text']: ins for ins in all_insights}.values()

    for ins in unique_insights:
        ws_insights.append([
            ins['thread_file'], ins['insight_text'], ins['category'],
            ins['applicability_to_our_notebook'], ins['expected_cv_delta_qualitative'],
            ins['concrete_code_change_if_applicable']
        ])

    # Sort insights for top 20
    def rank_insight(ins):
        score = 0
        if ins['expected_cv_delta_qualitative'] == 'HIGH': score += 10
        if ins['expected_cv_delta_qualitative'] == 'MED': score += 5
        if ins['applicability_to_our_notebook'] == 'YES': score += 10
        if ins['concrete_code_change_if_applicable'] != "None" and not ins['concrete_code_change_if_applicable'].startswith("Look for"):
            score += 5
        return score

    sorted_insights = sorted(unique_insights, key=rank_insight, reverse=True)
    top20 = sorted_insights[:20]

    for ins in top20:
        ws_top20.append([
            ins['thread_file'], ins['insight_text'], ins['category'],
            ins['applicability_to_our_notebook'], ins['expected_cv_delta_qualitative'],
            ins['concrete_code_change_if_applicable']
        ])

    unique_gotchas = {g['failure_mode']: g for g in all_gotchas}.values()
    for g in unique_gotchas:
        ws_gotchas.append([g['failure_mode'], g['fix']])

    wb.save('competitions/nvidia-nemotron-model-reasoning-challenge/analysis/discussions_extract.xlsx')
    print(f"Generated {len(unique_insights)} insights, {len(top20)} top actionable, {len(unique_gotchas)} gotchas.")

if __name__ == '__main__':
    main()
