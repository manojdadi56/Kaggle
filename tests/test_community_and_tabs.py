import pytest
import os
from openpyxl import load_workbook

XLSX_PATH = "competitions/nvidia-nemotron-model-reasoning-challenge/analysis/community_and_tabs.xlsx"

def test_xlsx_exists():
    assert os.path.exists(XLSX_PATH)

def test_xlsx_sheets():
    wb = load_workbook(XLSX_PATH)
    expected_sheets = ['CommunityCode', 'CompetitionFacts', 'TechniqueBacklog', 'OurAnalysisSummary']
    for sheet in expected_sheets:
        assert sheet in wb.sheetnames

def test_community_code_rows():
    wb = load_workbook(XLSX_PATH)
    sheet = wb['CommunityCode']
    # Subtract 1 for header
    row_count = sheet.max_row - 1
    assert row_count >= 30, f"Expected >= 30 CommunityCode rows, got {row_count}"

def test_competition_facts_rows():
    wb = load_workbook(XLSX_PATH)
    sheet = wb['CompetitionFacts']
    row_count = sheet.max_row - 1
    assert row_count >= 20, f"Expected >= 20 CompetitionFacts rows, got {row_count}"

def test_technique_backlog_rows():
    wb = load_workbook(XLSX_PATH)
    sheet = wb['TechniqueBacklog']
    row_count = sheet.max_row - 1
    assert row_count >= 9, f"Expected >= 9 TechniqueBacklog rows, got {row_count}"
