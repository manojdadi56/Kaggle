"""Read-only Excel dashboard — render the git-JSON ledger to one workbook.

This gives the human "everything in one Excel sheet" view (like the SDLC skill)
WITHOUT making Excel the source of truth (state.json/events.jsonl stay the truth).
Regenerated each tick / on demand: `python -m orchestrator.tools dashboard`.
"""
from __future__ import annotations

from pathlib import Path

from . import config


import re as _re
_ILLEGAL = _re.compile(r"[\x00-\x08\x0b-\x0c\x0e-\x1f]")


def _cell(v):
    """Coerce a value to something openpyxl will accept (no lists/dicts/control chars)."""
    if v is None or isinstance(v, (int, float, bool)):
        return v
    if isinstance(v, str):
        return _ILLEGAL.sub("", v)
    if isinstance(v, (list, tuple, set)):
        return _ILLEGAL.sub("", ", ".join(str(x) for x in v))
    if isinstance(v, dict):
        import json as _json
        return _ILLEGAL.sub("", _json.dumps(v, default=str))[:1000]
    return _ILLEGAL.sub("", str(v))


def _ws_from_rows(wb, title, headers, rows):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for r in rows:
        ws.append([_cell(r.get(h) if isinstance(r, dict) else r) for h in headers])
    return ws


def build_dashboard(state: dict, out_path: Path) -> Path:
    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)

    # Status overview
    ov = wb.create_sheet("Status")
    ov.append(["field", "value"])
    for k in ("active_competition", "updated_at", "best_cv"):
        ov.append([k, state.get(k)])
    ov.append(["in_flight_sessions", len(state.get("sessions", {}))])
    ov.append(["in_flight_gpu_runs", len(state.get("gpu_runs", {}))])
    ov.append(["tasks", len(state.get("tasks", {}))])
    ov.append(["hypotheses", len(state.get("hypotheses", {}))])
    ov.append(["experiments", len(state.get("experiments", {}))])
    for day, n in (state.get("submit_counter", {}) or {}).items():
        ov.append([f"submits {day}", n])

    def items(coll):
        return [{**v, "id": k} for k, v in state.get(coll, {}).items()]

    _ws_from_rows(wb, "Tasks", ["id", "title", "status", "hypothesis", "allowed_area", "story", "spec_path"], items("tasks"))
    _ws_from_rows(wb, "Hypotheses", ["id", "statement", "status", "expected_effect", "source_refs"], items("hypotheses"))
    _ws_from_rows(wb, "Experiments", ["id", "hypothesis", "backend", "cv_score", "status", "notes"], items("experiments"))
    _ws_from_rows(wb, "Suggestions", ["id", "title", "change_size", "status"], items("suggestions"))
    _ws_from_rows(wb, "Decisions", ["id", "decision", "rationale"], items("decisions"))
    _ws_from_rows(wb, "Metrics", ["name", "value", "unit", "at"], state.get("metrics", []))
    _ws_from_rows(wb, "Sessions", ["id", "task_id", "state", "pr_url"],
                  [{**v, "id": k} for k, v in state.get("sessions", {}).items()])

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def default_path() -> Path:
    return config.STATE_DIR / "dashboard.xlsx"
