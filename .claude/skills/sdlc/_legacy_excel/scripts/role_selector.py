#!/usr/bin/env python3
import argparse
import json
from pathlib import Path
from openpyxl import load_workbook

ROLES = ["planner", "owner", "reviewer", "tester", "maintainer", "innovator", "validator", "reporter"]


def records(wb, sheet):
    if sheet not in wb.sheetnames:
        return []
    ws = wb[sheet]
    headers = [c.value for c in ws[1]]
    rows = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        if not any(values):
            continue
        rows.append(dict(zip(headers, values)))
    return rows


def last_successful_role(wb):
    rows = records(wb, "RoleHistory")
    for row in reversed(rows):
        role = row.get("role_selected")
        if role:
            return str(role).lower()
    rows = records(wb, "Runs_Attendance")
    for row in reversed(rows):
        if str(row.get("status", "")).upper() in {"CLOSED", "COMPLETED", "PATCH_APPLIED"} and row.get("role_selected"):
            return str(row["role_selected"]).lower()
    return None


def status_count(rows, statuses):
    statuses = {s.upper() for s in statuses}
    return sum(1 for r in rows if str(r.get("status", "")).upper() in statuses)


def active_lock(rows, lock_type=None, scope_id=None):
    for r in rows:
        if str(r.get("status", "")).upper() != "ACTIVE":
            continue
        if lock_type and r.get("lock_type") != lock_type:
            continue
        if scope_id and r.get("scope_id") != scope_id:
            continue
        return True
    return False


def score_roles(wb, prevent_repeat=True):
    tasks = records(wb, "Tasks")
    suggestions = records(wb, "Suggestions")
    feedback = records(wb, "UserFeedback_Inbox")
    locks = records(wb, "Locks")
    innovation = records(wb, "InnovationLog")
    maint = records(wb, "MaintainerFindings")
    validations = records(wb, "Validations")
    previous = last_successful_role(wb)
    feasible = []
    rejected = []

    def add(role, score, reason):
        if prevent_repeat and previous == role:
            rejected.append({"role": role, "reason": "previous successful role"})
            return
        feasible.append({"role": role, "score": score, "reason": reason})

    if status_count(feedback, {"", "NEW", "OPEN", "READY"}) or status_count(suggestions, {"PROPOSED", "VALIDATION_REQUIRED", "NEEDS_INFO"}) or not tasks:
        add("planner", 80, "feedback, suggestions, validation results, or empty backlog need planning")
    else:
        rejected.append({"role": "planner", "reason": "no planning input"})

    if status_count(tasks, {"READY", "SATURATED", "CHANGES_REQUESTED"}) and not active_lock(locks, "owner_project_lock"):
        add("owner", 90, "ready or resumable owner task exists and no owner project lock is active")
    else:
        rejected.append({"role": "owner", "reason": "no ready task or owner lock active"})

    if status_count(tasks, {"READY_FOR_REVIEW", "DONE_PENDING_REVIEW"}):
        add("reviewer", 75, "review-ready task exists")
    else:
        rejected.append({"role": "reviewer", "reason": "no review-ready task"})

    if status_count(tasks, {"READY_FOR_TEST", "DONE_PENDING_TEST"}):
        add("tester", 72, "implementation exists and test gate is pending")
    else:
        rejected.append({"role": "tester", "reason": "no implementation awaiting tests"})

    if status_count(tasks, {"DONE", "READY_FOR_REVIEW", "READY_FOR_TEST"}) and len(maint) < max(1, len(tasks) // 4):
        add("maintainer", 55, "completed or reviewable work exists and maintenance scan is useful")
    else:
        rejected.append({"role": "maintainer", "reason": "no useful maintenance trigger"})

    if len(innovation) == 0 or status_count(suggestions, {"NEEDS_INFO"}):
        add("innovator", 45, "research log is empty or planner needs external information")
    else:
        rejected.append({"role": "innovator", "reason": "no active research trigger"})

    pending_validations = [s for s in suggestions if str(s.get("status", "")).upper() in {"VALIDATION_REQUIRED", "VALIDATING"}]
    if pending_validations:
        add("validator", 78, "suggestions are awaiting validation")
    else:
        rejected.append({"role": "validator", "reason": "no pending validation"})

    add("reporter", 10, "fallback summary role")
    feasible.sort(key=lambda x: x["score"], reverse=True)
    return {"previous_role": previous, "selected": feasible[0] if feasible else None, "feasible": feasible, "rejected": rejected}


def main():
    parser = argparse.ArgumentParser(description="Score feasible SDLC roles from a project workbook.")
    parser.add_argument("--workbook", required=True)
    parser.add_argument("--allow-repeat", action="store_true")
    args = parser.parse_args()
    wb = load_workbook(args.workbook, data_only=True)
    print(json.dumps(score_roles(wb, prevent_repeat=not args.allow_repeat), indent=2))


if __name__ == "__main__":
    main()
