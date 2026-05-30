#!/usr/bin/env python3
"""Rebuild projection sheets from the Events sheet of a state workbook.

Use when a projection sheet (Tasks, Suggestions, etc.) is corrupted, or to
verify the Events stream is internally consistent.

Strategy:
  1. Snapshot the workbook before any change.
  2. Read Events in chronological order.
  3. Re-derive each projection by replaying ops:
     - create_*  -> append a new row to the target sheet
     - update_*  -> mutate the row by entity_id
     - register_artifact -> append to ArtifactRegistry
     - add_metric / add_memory_candidate -> append to respective sheet
  4. Recompute Dashboard formulas (handled by openpyxl on open in Excel).

Usage:
  python3 workbook_recover.py --workbook PATH [--out PATH] [--dry-run]
"""
import argparse
import json
import shutil
import sys
from datetime import datetime, timezone
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print(json.dumps({"error": "openpyxl_not_installed"}), file=sys.stderr)
    sys.exit(2)

# Lazy import of SHEETS so this script lives in the same dir as workbook_schema.py
sys.path.insert(0, str(Path(__file__).parent))
from workbook_schema import SHEETS  # noqa: E402

PROJECTIONS = [s for s in SHEETS.keys() if s not in ("Events", "Dashboard", "UserContext", "UserFeedback_Inbox")]


def headers(ws):
    return [c.value for c in ws[1]]


def event_rows(wb):
    if "Events" not in wb.sheetnames:
        return []
    ws = wb["Events"]
    h = headers(ws)
    out = []
    for r in range(2, ws.max_row + 1):
        row = {k: ws.cell(r, i + 1).value for i, k in enumerate(h) if k}
        if any(v not in (None, "") for v in row.values()):
            out.append(row)
    out.sort(key=lambda x: str(x.get("created_at", "")))
    return out


def reset_projection(wb, sheet_name):
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    ws.append(SHEETS[sheet_name])


def replay_event(wb, ev):
    et = (ev.get("event_type") or "").upper()
    sheet = ev.get("entity_type") or ""
    # Map a few common event_types -> sheet
    sheet_map = {
        "CREATE_TASK": "Tasks",
        "CREATE_SUPPORT_WORK": "SupportWork",
        "CREATE_SUGGESTION": "Suggestions",
        "CREATE_VALIDATION": "Validations",
        "CREATE_DECISION": "Decisions",
        "REGISTER_ARTIFACT": "ArtifactRegistry",
        "ADD_METRIC": "Metrics",
        "ADD_MEMORY_CANDIDATE": "ProjectMemory_Index",
        "START_ATTENDANCE": "Runs_Attendance",
        "CLOSE_ATTENDANCE": "Runs_Attendance",
        "UPDATE_STATUS": None,
        "UPDATE_ROW_BY_ID": None,
    }
    target = sheet_map.get(et)
    details_raw = ev.get("details_json")
    details = {}
    if details_raw:
        try:
            details = json.loads(details_raw)
        except Exception:
            details = {}
    data = details.get("data") if isinstance(details, dict) else {}
    if et.startswith("CREATE_") or et in ("REGISTER_ARTIFACT", "ADD_METRIC", "ADD_MEMORY_CANDIDATE"):
        if not target:
            return
        ws = wb[target]
        h = headers(ws)
        ws.append([data.get(k) if isinstance(data, dict) else "" for k in h])
    elif et in ("UPDATE_STATUS", "UPDATE_ROW_BY_ID"):
        target = details.get("sheet") if isinstance(details, dict) else None
        if not target or target not in wb.sheetnames:
            return
        ws = wb[target]
        h = headers(ws)
        eid = ev.get("entity_id")
        # find row
        id_field = next((f for f in ("task_id", "story_id", "phase_id", "suggestion_id", "validation_id", "decision_id", "support_id", "artifact_id", "metric_id") if f in h), None)
        if not id_field:
            return
        col = h.index(id_field) + 1
        for r in range(2, ws.max_row + 1):
            if str(ws.cell(r, col).value) == str(eid):
                if et == "UPDATE_STATUS" and "status" in h:
                    ws.cell(r, h.index("status") + 1, details.get("to_status"))
                else:
                    payload = details.get("data", {}) if isinstance(details, dict) else {}
                    for k, v in payload.items():
                        if k in h:
                            ws.cell(r, h.index(k) + 1, v)
                break
    elif et in ("START_ATTENDANCE", "CLOSE_ATTENDANCE"):
        ws = wb["Runs_Attendance"]
        h = headers(ws)
        run_id = ev.get("run_id")
        if "run_id" not in h or not run_id:
            return
        run_col = h.index("run_id") + 1
        payload = dict(data) if isinstance(data, dict) else {}
        # Always carry the event-level run_id and project_id into the projection row
        payload.setdefault("run_id", run_id)
        if "project_id" in h and ev.get("project_id"):
            payload.setdefault("project_id", ev.get("project_id"))
        # Locate or create the row keyed by run_id
        target_row = None
        for r in range(2, ws.max_row + 1):
            if str(ws.cell(r, run_col).value) == str(run_id):
                target_row = r
                break
        if target_row is None:
            ws.append([payload.get(k, "") for k in h])
        else:
            for k, v in payload.items():
                if k in h:
                    ws.cell(target_row, h.index(k) + 1, v)


def recover(workbook_path, out_path, dry_run=False):
    wb = load_workbook(workbook_path)
    events = event_rows(wb)
    if not dry_run:
        # Reset projections
        for s in PROJECTIONS:
            reset_projection(wb, s)
        # Replay events
        for ev in events:
            replay_event(wb, ev)
        wb.save(out_path)
    return {"event_count": len(events), "projections_reset": PROJECTIONS, "out": str(out_path), "dry_run": dry_run}


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--workbook", required=True)
    p.add_argument("--out", default=None, help="If omitted, overwrites workbook in place after backup.")
    p.add_argument("--dry-run", action="store_true")
    args = p.parse_args()
    wb_path = Path(args.workbook)
    if not wb_path.exists():
        print(json.dumps({"error": "workbook_not_found", "path": str(wb_path)}))
        sys.exit(1)
    out = Path(args.out) if args.out else wb_path
    if not args.dry_run and out == wb_path:
        backup = wb_path.with_suffix(f".backup-{datetime.now(timezone.utc).strftime('%Y%m%d-%H%M%S')}.xlsx")
        shutil.copy2(wb_path, backup)
        print(json.dumps({"backup": str(backup)}), file=sys.stderr)
    result = recover(wb_path, out, dry_run=args.dry_run)
    print(json.dumps(result, indent=2))


if __name__ == "__main__":
    main()
