#!/usr/bin/env python3
import argparse
from pathlib import Path
from datetime import datetime, timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from workbook_schema import SHEETS

HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
TITLE_FILL = PatternFill("solid", fgColor="EEF5FB")
THIN = Side(style="thin", color="D9E2EC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def now_iso():
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def style_header(ws, row=1):
    for cell in ws[row]:
        cell.font = Font(bold=True, color="1F2937")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.freeze_panes = "A2"


def autosize(ws):
    for col in ws.columns:
        max_len = 8
        letter = get_column_letter(col[0].column)
        for cell in col[:60]:
            val = cell.value
            if val is not None:
                max_len = max(max_len, min(len(str(val)), 60))
        ws.column_dimensions[letter].width = min(max_len + 2, 42)


def create_state_workbook(path: str, project_id: str = "PROJECT"):
    wb = Workbook()
    default = wb.active
    wb.remove(default)
    for sheet_name, headers in SHEETS.items():
        ws = wb.create_sheet(sheet_name)
        ws.append(headers)
        style_header(ws)
        if sheet_name == "Dashboard":
            ws.append(["project_id", project_id, "Set by bootstrap or project.yaml"])
            ws.append(["created_at", now_iso(), "UTC"])
            ws.append(["ready_tasks", '=COUNTIF(Tasks!G:G,"READY")', "Owner-feasible tasks"])
            ws.append(["in_progress_tasks", '=COUNTIF(Tasks!G:G,"IN_PROGRESS")+COUNTIF(Tasks!G:G,"SATURATED")', "Active or resumable tasks"])
            ws.append(["done_tasks", '=COUNTIF(Tasks!G:G,"DONE")', "Completed tasks"])
            ws.append(["open_suggestions", '=COUNTIF(Suggestions!I:I,"PROPOSED")+COUNTIF(Suggestions!I:I,"VALIDATION_REQUIRED")', "Needs planner/validation"])
            ws.append(["last_updated", now_iso(), "Updated by state-ledger"])
            for row in ws.iter_rows(min_row=2, max_row=8):
                for cell in row:
                    cell.border = BORDER
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
            ws["A1"].fill = TITLE_FILL
        autosize(ws)
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def create_index_workbook(path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Projects"
    sheets = {
        "Projects": ["project_id", "project_name", "project_root", "central_root", "state_workbook", "status", "cadence", "current_phase", "created_at", "updated_at", "notes"],
        "GlobalRuns": ["run_id", "project_id", "started_at", "ended_at", "role_selected", "status", "summary", "next_recommended_role"],
        "SkillConfig": ["key", "value", "scope", "notes"],
        "KnownUsers": ["user_id", "display_name", "role", "contact", "notes"],
        "CentralEvents": ["event_id", "project_id", "run_id", "created_at", "event_type", "summary", "details_json"],
        "Dashboard": ["field", "value", "notes"]
    }
    for i, (name, headers) in enumerate(sheets.items()):
        if i == 0:
            ws = wb["Projects"]
        else:
            ws = wb.create_sheet(name)
        ws.append(headers)
        style_header(ws)
        if name == "Dashboard":
            ws.append(["created_at", now_iso(), "UTC"])
            ws.append(["project_count", "=COUNTA(Projects!A:A)-1", "Registered projects"])
            ws.append(["open_projects", '=COUNTIF(Projects!F:F,"ACTIVE")', "Active projects"])
        autosize(ws)
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def main():
    parser = argparse.ArgumentParser(description="Create SDLC state or index workbooks.")
    parser.add_argument("--type", choices=["state", "index"], required=True)
    parser.add_argument("--path", required=True)
    parser.add_argument("--project-id", default="PROJECT")
    args = parser.parse_args()
    if args.type == "state":
        create_state_workbook(args.path, args.project_id)
    else:
        create_index_workbook(args.path)
    print(args.path)


if __name__ == "__main__":
    main()
