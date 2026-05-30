#!/usr/bin/env python3
"""Read a state.xlsx workbook and emit a JSON snapshot for read-only consumption.

Usage:
  python3 load_project_state.py --workbook PATH [--limit-events N] [--out FILE]

The snapshot includes every sheet's headers + rows (capped per sheet) so that
the role stage can reason about state without re-reading the workbook.
"""
import argparse
import json
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print(json.dumps({"error": "openpyxl_not_installed", "hint": "pip install openpyxl"}), file=sys.stderr)
    sys.exit(2)


def headers(ws):
    return [c.value for c in ws[1]]


def rows(ws, limit=None):
    h = headers(ws)
    out = []
    max_row = ws.max_row
    if limit is not None and max_row > 1 + limit:
        # keep the most recent rows for sheets like Events, RoleHistory
        start = max_row - limit + 1
    else:
        start = 2
    for r in range(start, max_row + 1):
        row = {}
        for i, key in enumerate(h):
            if key is None:
                continue
            row[key] = ws.cell(r, i + 1).value
        if any(v not in (None, "") for v in row.values()):
            out.append(row)
    return out


def snapshot(workbook_path, limit_events=200):
    wb = load_workbook(workbook_path, data_only=True, read_only=False)
    out = {"workbook": str(workbook_path), "sheets": {}}
    big_sheets = {"Events", "RoleHistory", "Runs_Attendance"}
    for name in wb.sheetnames:
        ws = wb[name]
        h = headers(ws)
        limit = limit_events if name in big_sheets else None
        out["sheets"][name] = {
            "headers": h,
            "row_count": max(ws.max_row - 1, 0),
            "rows": rows(ws, limit=limit),
        }
    return out


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--workbook", required=True)
    p.add_argument("--limit-events", type=int, default=200)
    p.add_argument("--out", default=None)
    args = p.parse_args()
    wb_path = Path(args.workbook)
    if not wb_path.exists():
        print(json.dumps({"error": "workbook_not_found", "path": str(wb_path)}))
        sys.exit(1)
    data = snapshot(wb_path, limit_events=args.limit_events)
    payload = json.dumps(data, indent=2, default=str)
    if args.out:
        Path(args.out).write_text(payload, encoding="utf-8")
        print(json.dumps({"out": args.out, "sheet_count": len(data["sheets"])}))
    else:
        print(payload)


if __name__ == "__main__":
    main()
