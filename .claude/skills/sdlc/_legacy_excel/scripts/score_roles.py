#!/usr/bin/env python3
"""Score and rank feasible roles for the next SDLC cycle.

Reads the workbook, applies feasibility + adjacency rules, and returns a
deterministic ranked list. Tie-breaks alphabetically.

Usage:
  python3 score_roles.py --workbook PATH [--previous-role ROLE] [--override ROLE]

Output: JSON {"ranked": [{"role": ..., "score": ..., "reason": ...}], "feasible": [...], "rejected": [{"role": ..., "reason": ...}]}
"""
import argparse
import json
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print(json.dumps({"error": "openpyxl_not_installed"}), file=sys.stderr)
    sys.exit(2)

ALL_ROLES = ["planner", "owner", "reviewer", "tester", "maintainer", "innovator", "validator", "reporter"]


def headers(ws):
    return [c.value for c in ws[1]]


def rows_as_dicts(ws):
    h = headers(ws)
    return [
        {k: ws.cell(r, i + 1).value for i, k in enumerate(h) if k}
        for r in range(2, ws.max_row + 1)
    ]


def feasibility(state):
    """Compute per-role feasibility + base score from current workbook state."""
    tasks = state.get("Tasks", [])
    suggestions = state.get("Suggestions", [])
    validations = state.get("Validations", [])
    feedback = state.get("UserFeedback_Inbox", [])
    review_findings = state.get("ReviewFindings", [])
    test_findings = state.get("TestFindings", [])
    out = {}

    ready_tasks = [t for t in tasks if t.get("status") == "READY"]
    in_progress = [t for t in tasks if t.get("status") in ("IN_PROGRESS", "SATURATED", "CLAIMED")]
    review_queue = [t for t in tasks if t.get("status") == "READY_FOR_REVIEW"]
    test_queue = [t for t in tasks if t.get("status") == "READY_FOR_TEST"]
    blocked_tasks = [t for t in tasks if t.get("status") == "BLOCKED"]
    open_suggestions = [s for s in suggestions if s.get("status") in ("PROPOSED", "VALIDATION_REQUIRED")]
    pending_feedback = [f for f in feedback if f.get("status") in (None, "", "NEW", "OPEN")]
    open_review_findings = [f for f in review_findings if f.get("status") in (None, "", "OPEN", "NEW")]

    out["planner"] = {
        "feasible": bool(pending_feedback or open_suggestions or not ready_tasks or blocked_tasks),
        "score": 5 * len(pending_feedback) + 3 * len(open_suggestions) + (4 if not ready_tasks and not in_progress else 0) + 2 * len(blocked_tasks),
        "reason": f"{len(pending_feedback)} feedback, {len(open_suggestions)} suggestions, {len(ready_tasks)} ready tasks, {len(blocked_tasks)} blocked",
    }
    out["owner"] = {
        "feasible": bool(ready_tasks or in_progress),
        "score": 8 * len(ready_tasks) + 6 * len(in_progress),
        "reason": f"{len(ready_tasks)} ready tasks, {len(in_progress)} resumable",
    }
    out["reviewer"] = {
        "feasible": bool(review_queue),
        "score": 7 * len(review_queue),
        "reason": f"{len(review_queue)} tasks awaiting review",
    }
    out["tester"] = {
        "feasible": bool(test_queue),
        "score": 6 * len(test_queue) + 2 * len(open_review_findings),
        "reason": f"{len(test_queue)} tasks awaiting test, {len(open_review_findings)} findings",
    }
    out["maintainer"] = {
        "feasible": True,
        "score": 1 + len(test_findings) // 5,
        "reason": "always-feasible (low-priority floor)",
    }
    out["innovator"] = {
        "feasible": True,
        "score": 1,
        "reason": "always-feasible (low-priority floor)",
    }
    open_validations = [v for v in suggestions if v.get("status") == "VALIDATION_REQUIRED"]
    out["validator"] = {
        "feasible": bool(open_validations),
        "score": 5 * len(open_validations),
        "reason": f"{len(open_validations)} suggestions awaiting validation",
    }
    out["reporter"] = {
        "feasible": True,
        "score": 0,
        "reason": "fallback when nothing else fits",
    }
    return out


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--workbook", required=True)
    p.add_argument("--previous-role", default=None)
    p.add_argument("--override", default=None, help="Force a specific role; ignores adjacency.")
    args = p.parse_args()

    wb = load_workbook(args.workbook, data_only=True)
    state = {name: rows_as_dicts(wb[name]) for name in wb.sheetnames}
    feas = feasibility(state)

    rejected = []
    candidates = []
    for role in ALL_ROLES:
        info = feas[role]
        if args.override and role != args.override:
            rejected.append({"role": role, "reason": "explicit_override_excluded"})
            continue
        if not info["feasible"]:
            rejected.append({"role": role, "reason": f"infeasible: {info['reason']}"})
            continue
        if not args.override and args.previous_role and role == args.previous_role:
            rejected.append({"role": role, "reason": "adjacent_repeat"})
            continue
        candidates.append({"role": role, "score": info["score"], "reason": info["reason"]})

    candidates.sort(key=lambda x: (-x["score"], x["role"]))
    out = {
        "ranked": candidates,
        "rejected": rejected,
        "feasible": [c["role"] for c in candidates],
        "selected": candidates[0]["role"] if candidates else None,
    }
    print(json.dumps(out, indent=2, default=str))


if __name__ == "__main__":
    main()
