#!/usr/bin/env python3
import argparse
import json
import os
import shutil
import tempfile
from pathlib import Path
from datetime import datetime, timezone
from openpyxl import load_workbook
from workbook_schema import SHEETS
from workbook_init import create_state_workbook


def now_iso():
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def ensure_workbook(path, project_id="PROJECT"):
    path = Path(path)
    if not path.exists():
        create_state_workbook(str(path), project_id)
    wb = load_workbook(path)
    for sheet, headers in SHEETS.items():
        if sheet not in wb.sheetnames:
            ws = wb.create_sheet(sheet)
            ws.append(headers)
        else:
            ws = wb[sheet]
            existing = [c.value for c in ws[1]]
            if not any(existing):
                ws.append(headers)
    return wb


def headers(ws):
    return [c.value for c in ws[1]]


def row_to_dict(ws, row):
    return dict(zip(headers(ws), [c.value for c in row]))


def append_row(ws, data):
    h = headers(ws)
    ws.append([data.get(k, "") for k in h])


def find_row(ws, field, value):
    h = headers(ws)
    if field not in h:
        raise ValueError(f"Field {field} not found in sheet {ws.title}")
    idx = h.index(field) + 1
    for row_idx in range(2, ws.max_row + 1):
        if str(ws.cell(row_idx, idx).value) == str(value):
            return row_idx
    return None


def set_fields(ws, row_idx, data):
    h = headers(ws)
    for key, value in data.items():
        if key not in h:
            continue
        ws.cell(row_idx, h.index(key) + 1, value)
    if "updated_at" in h:
        ws.cell(row_idx, h.index("updated_at") + 1, now_iso())
    if "version" in h:
        col = h.index("version") + 1
        old = ws.cell(row_idx, col).value or 0
        try:
            old = int(old)
        except Exception:
            old = 0
        ws.cell(row_idx, col, old + 1)


def operation_seen(wb, key):
    if not key or "Events" not in wb.sheetnames:
        return False
    ws = wb["Events"]
    h = headers(ws)
    if "idempotency_key" not in h:
        return False
    col = h.index("idempotency_key") + 1
    for row_idx in range(2, ws.max_row + 1):
        if ws.cell(row_idx, col).value == key:
            return True
    return False


def append_event(wb, patch, op):
    ws = wb["Events"]
    data = dict(op.get("data", {}))
    data.setdefault("event_id", f"EVT-{patch['run_id']}-{ws.max_row:05d}")
    data.setdefault("project_id", patch.get("project_id"))
    data.setdefault("run_id", patch.get("run_id"))
    data.setdefault("created_at", now_iso())
    data.setdefault("idempotency_key", op.get("idempotency_key", ""))
    if "details_json" not in data:
        details = {k: v for k, v in op.items() if k not in {"op", "data"}}
        data["details_json"] = json.dumps(details, sort_keys=True)
    append_row(ws, data)


def apply_operation(wb, patch, op):
    key = op.get("idempotency_key")
    if operation_seen(wb, key):
        return "skipped-idempotent"
    name = op["op"]
    if name == "append_event":
        append_event(wb, patch, op)
        return "ok"
    if name in {"append_row", "create_task", "create_support_work", "create_suggestion", "create_validation", "create_decision", "register_artifact", "add_metric", "add_memory_candidate"}:
        sheet = op.get("sheet") or {
            "create_task": "Tasks",
            "create_support_work": "SupportWork",
            "create_suggestion": "Suggestions",
            "create_validation": "Validations",
            "create_decision": "Decisions",
            "register_artifact": "ArtifactRegistry",
            "add_metric": "Metrics",
            "add_memory_candidate": "ProjectMemory_Index"
        }.get(name)
        if not sheet:
            raise ValueError(f"No sheet provided for {name}")
        ws = wb[sheet]
        data = dict(op.get("data", {}))
        data.setdefault("project_id", patch.get("project_id"))
        data.setdefault("created_at", now_iso())
        data.setdefault("updated_at", now_iso())
        append_row(ws, data)
        append_event(wb, patch, {"op": "append_event", "idempotency_key": key, "data": {"event_type": name.upper(), "entity_type": sheet, "entity_id": data.get("task_id") or data.get("suggestion_id") or data.get("artifact_id") or "", "summary": data.get("title") or data.get("summary") or name}})
        return "ok"
    if name in {"update_row_by_id", "update_status"}:
        sheet = op["sheet"]
        ws = wb[sheet]
        field = op.get("entity_id_field") or ("task_id" if sheet == "Tasks" else "suggestion_id")
        row_idx = find_row(ws, field, op["entity_id"])
        if row_idx is None:
            raise ValueError(f"No row found in {sheet} where {field}={op['entity_id']}")
        if name == "update_status":
            h = headers(ws)
            if "status" not in h:
                raise ValueError(f"Sheet {sheet} has no status field")
            status_col = h.index("status") + 1
            current = ws.cell(row_idx, status_col).value
            expected = op.get("from_status")
            if expected and str(current) != str(expected):
                raise ValueError(f"Status conflict for {op['entity_id']}: expected {expected}, found {current}")
            set_fields(ws, row_idx, {"status": op["to_status"]})
        else:
            set_fields(ws, row_idx, op.get("data", {}))
        append_event(wb, patch, {"op": "append_event", "idempotency_key": key, "data": {"event_type": name.upper(), "entity_type": sheet, "entity_id": op["entity_id"], "summary": op.get("summary", name)}})
        return "ok"
    if name in {"start_attendance", "close_attendance"}:
        ws = wb["Runs_Attendance"]
        run_id = op.get("run_id", patch.get("run_id"))
        row_idx = find_row(ws, "run_id", run_id)
        data = dict(op.get("data", {}))
        data.setdefault("run_id", run_id)
        data.setdefault("project_id", patch.get("project_id"))
        if name == "start_attendance":
            data.setdefault("started_at", now_iso())
            data.setdefault("status", "STARTED")
            if row_idx is None:
                append_row(ws, data)
            else:
                set_fields(ws, row_idx, data)
        else:
            data.setdefault("ended_at", now_iso())
            data.setdefault("status", "CLOSED")
            if row_idx is None:
                append_row(ws, data)
            else:
                set_fields(ws, row_idx, data)
        append_event(wb, patch, {"op": "append_event", "idempotency_key": key, "data": {"event_type": name.upper(), "entity_type": "run", "entity_id": run_id, "summary": data.get("high_level_completed") or data.get("status") or name}})
        return "ok"
    raise ValueError(f"Unsupported operation: {name}")


def atomic_save(wb, path):
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False, dir=str(path.parent)) as tmp:
        tmp_path = Path(tmp.name)
    try:
        wb.save(tmp_path)
        os.replace(tmp_path, path)
    finally:
        if tmp_path.exists():
            tmp_path.unlink(missing_ok=True)


def apply_patch(workbook, patch_file):
    patch = json.loads(Path(patch_file).read_text(encoding="utf-8"))
    wb = ensure_workbook(workbook, patch.get("project_id", "PROJECT"))
    results = []
    for op in patch.get("operations", []):
        results.append({"op": op.get("op"), "result": apply_operation(wb, patch, op)})
    atomic_save(wb, workbook)
    return results


def snapshot(workbook, out_dir):
    out = Path(out_dir)
    out.mkdir(parents=True, exist_ok=True)
    name = f"{Path(workbook).stem}-{datetime.now(timezone.utc).strftime('%Y%m%d-%H%M%S')}.xlsx"
    target = out / name
    shutil.copy2(workbook, target)
    return target


def main():
    parser = argparse.ArgumentParser(description="Apply SDLC StatePatch operations to a workbook.")
    sub = parser.add_subparsers(dest="cmd", required=True)
    p_init = sub.add_parser("init")
    p_init.add_argument("--workbook", required=True)
    p_init.add_argument("--project-id", default="PROJECT")
    p_apply = sub.add_parser("apply-patch")
    p_apply.add_argument("--workbook", required=True)
    p_apply.add_argument("--patch", required=True)
    p_snap = sub.add_parser("snapshot")
    p_snap.add_argument("--workbook", required=True)
    p_snap.add_argument("--out-dir", required=True)
    args = parser.parse_args()
    if args.cmd == "init":
        create_state_workbook(args.workbook, args.project_id)
        print(json.dumps({"workbook": args.workbook, "status": "created"}))
    elif args.cmd == "apply-patch":
        print(json.dumps({"results": apply_patch(args.workbook, args.patch)}, indent=2))
    elif args.cmd == "snapshot":
        print(json.dumps({"snapshot": str(snapshot(args.workbook, args.out_dir))}))


if __name__ == "__main__":
    main()
